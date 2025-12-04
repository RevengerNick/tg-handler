import asyncio
import os
import logging
import re
import time
import json
import psutil
import struct # Для создания WAV заголовков
import mimetypes
from io import BytesIO
from datetime import datetime

# Pyrogram
from pyrogram import Client, filters, idle
from pyrogram.errors import (
    SessionPasswordNeeded,
    PhoneCodeInvalid,
    PasswordHashInvalid,
    PhoneCodeExpired
)

# Utils
import aiohttp
import markdown
import yt_dlp
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from yandex_music import Client as YMClient
from telegraph import Telegraph
from dotenv import load_dotenv

# Google Gen AI
from google import genai
from google.genai import types

# --- SETUP LOGGING ---
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.ERROR
)
logger = logging.getLogger(__name__)

AVAILABLE_VOICES = {
    # --- Мужские ---
    "1": {"name": "Puck", "gender": "M", "desc": "Бодрый, средний тон"},
    "2": {"name": "Charon", "gender": "M", "desc": "Глубокий, низкий"},
    "3": {"name": "Fenrir", "gender": "M", "desc": "Басистый, энергичный"},
    "4": {"name": "Orus", "gender": "M", "desc": "Твердый, ниже среднего"},
    "5": {"name": "Enceladus", "gender": "M", "desc": "С придыханием, низкий"},
    "6": {"name": "Iapetus", "gender": "M", "desc": "Чистый, ниже среднего"},
    "7": {"name": "Umbriel", "gender": "M", "desc": "Спокойный, ниже среднего"},
    "8": {"name": "Algieba", "gender": "M", "desc": "Гладкий, низкий"},
    "9": {"name": "Algenib", "gender": "M", "desc": "Хриплый, низкий"},
    "10": {"name": "Achernar", "gender": "M", "desc": "Мягкий, высокий"},
    "11": {"name": "Alnilam", "gender": "M", "desc": "Твердый, ниже среднего"},
    "12": {"name": "Schedar", "gender": "M", "desc": "Ровный, ниже среднего"},
    "13": {"name": "Zubenelgenubi", "gender": "M", "desc": "Небрежный, ниже среднего"},

    # --- Женские ---
    "14": {"name": "Zephyr", "gender": "F", "desc": "Светлый, высокий"},
    "20": {"name": "Despina", "gender": "F", "desc": "Гладкий, средний"},
    "21": {"name": "Erinome", "gender": "F", "desc": "Чистый, средний"},
    "30": {"name": "Sulafat", "gender": "F", "desc": "Теплый, средний"},
    "15": {"name": "Kore", "gender": "F", "desc": "Твердый, средний"},
    "16": {"name": "Leda", "gender": "F", "desc": "Молодой, высокий"},
    "17": {"name": "Aoede", "gender": "F", "desc": "Легкий, средний"},
    "18": {"name": "Callirrhoe", "gender": "F", "desc": "Беззаботный, средний"},
    "19": {"name": "Autonoe", "gender": "F", "desc": "Яркий, средний"},
    "22": {"name": "Rasalgethi", "gender": "F", "desc": "Информативный, средний"},
    "23": {"name": "Laomedeia", "gender": "F", "desc": "Бодрый, высокий"},
    "24": {"name": "Gacrux", "gender": "F", "desc": "Зрелый, средний"},
    "25": {"name": "Pulcherrima", "gender": "F", "desc": "Прямолинейный, средний"},
    "26": {"name": "Achird", "gender": "F", "desc": "Дружелюбный, ниже среднего"},
    "27": {"name": "Vindemiatrix", "gender": "F", "desc": "Нежный, средний"},
    "28": {"name": "Sadachbia", "gender": "F", "desc": "Живой, низкий"},
    "29": {"name": "Sadaltager", "gender": "F", "desc": "Знающий, средний"}
}

# Обновленный список для диалогов (чередуем разные тембры для контраста)
VOICE_NAMES_LIST = [
    "Puck", "Zephyr", "Fenrir", "Leda", "Charon", "Aoede",
    "Orus", "Autonoe", "Algenib", "Erinome", "Enceladus", "Kore"
]
AVAILABLE_TTS_MODELS = {
    "1": "gemini-2.5-pro-preview-tts", # PRO (Лучшее качество)
    "2": "gemini-2.5-flash-preview-tts",           # FLASH (Быстрее)
}
load_dotenv()

# --- CONFIGURATION ---
API_ID = int(os.getenv("API_ID", 0))
API_HASH = os.getenv("API_HASH", "")
PHONES = os.getenv("PHONES", "").split(",")
GEMINI_KEY = os.getenv("GEMINI_API_KEY")
YANDEX_TOKEN = os.getenv("YANDEX_TOKEN")
EXCHANGE_KEY = os.getenv("EXCHANGE_API_KEY")
SETTINGS_FILE = "settings.json"

AVAILABLE_MODELS = {
    "1": {"id": "gemini-2.5-flash", "name": "⚡️ 2.5 Flash (Google Search)", "search": True},
    "2": {"id": "gemini-2.5-pro", "name": "🧠 2.5 Pro (Thinking)", "search": False},
    "3": {"id": "gemini-2.0-flash", "name": "🚀 2.0 Flash (Fast)", "search": False},
}

# --- GLOBAL STATE ---
# Хранилище АСИНХРОННЫХ сессий чата: {chat_id: chat_session_object}
ASYNC_CHAT_SESSIONS = {}

SETTINGS = {
    "model_key": "1",      # Текстовая модель
    "voice_key": "1",      # Голос (Puck по дефолту)
    "tts_model_key": "1",  # Модель озвучки (Pro по дефолту)
    "sys_global": "",
    "sys_chats": {}
}
# Telegraph Init
telegraph_client = Telegraph()
try:
    telegraph_client.create_account(short_name='GeminiBot')
except: pass

def load_settings():
    global SETTINGS
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                saved = json.load(f)
                for k, v in saved.items(): SETTINGS[k] = v
            model_info = AVAILABLE_MODELS.get(SETTINGS.get("model_key", "1"))
            print(f"⚙️ Config Loaded. Model: {model_info['name']}")
        except Exception as e: print(f"⚠️ Config Err: {e}")

def save_settings():
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(SETTINGS, f, indent=4, ensure_ascii=False)
    except: pass

load_settings()

# --- INIT CLIENTS ---
ai_client = None
if GEMINI_KEY:
    try:
        ai_client = genai.Client(api_key=GEMINI_KEY)
    except Exception as e:
        print(f"❌ Ошибка Init Gemini: {e}")

ym_client = YMClient(YANDEX_TOKEN).init() if YANDEX_TOKEN else None

# --- HELPER FUNCTIONS ---

async def add_stress_via_gemini(text):
    """
    Просит Gemini расставить ударения для TTS.
    """
    if not ai_client: return text

    # Промпт жесткий, чтобы он вернул ТОЛЬКО текст без "Конечно, вот текст:"
    prompt = (
        "Расставь ударения в этом тексте, используя символ '́' (U+0301) ПОСЛЕ ударной гласной. "
        "Исправляй омографы по контексту. "
        "Верни ТОЛЬКО обработанный текст, без кавычек и вступлений.\n"
        f"Текст: {text}"
    )

    try:
        # Используем быстрый Flash для скорости
        model_id = "gemini-2.5-pro"
        response = await ai_client.aio.models.generate_content(
            model=model_id,
            contents=prompt
        )
        result = response.text.strip()
        return result
    except Exception as e:
        print(f"Stress Error: {e}")
        return text  # Если ошибка, возвращаем оригинал


def parse_audio_mime_type(mime_type: str):
    """Парсит частоту дискретизации и битность из MIME-типа."""
    bits_per_sample = 16
    rate = 24000
    parts = mime_type.split(";")
    for param in parts:
        param = param.strip()
        if param.lower().startswith("rate="):
            try:
                rate = int(param.split("=", 1)[1])
            except:
                pass
        elif param.startswith("audio/L"):
            try:
                bits_per_sample = int(param.split("L", 1)[1])
            except:
                pass
    return {"bits_per_sample": bits_per_sample, "rate": rate}


async def generate_multispeaker_tts(script_text, custom_cast=None):
    """
    Генерация диалога с несколькими спикерами.
    script_text: Текст сценария.
    custom_cast: Словарь { "ИмяСпикера": "ИмяГолоса" } (опционально).
    """
    if not ai_client: return None

    # 1. Поиск уникальных спикеров в тексте
    # Ищем любые "Имя:" или "1:" в начале строки
    speaker_pattern = re.compile(r"^([A-Za-zА-Яа-я0-9_ ]+):", re.MULTILINE)
    found_speakers = list(set(speaker_pattern.findall(script_text)))

    if not found_speakers:
        # Если формат не найден, считаем это монологом
        found_speakers = ["Narrator"]
        script_text = f"Narrator: {script_text}"

    print(f"DEBUG: Speakers found: {found_speakers}")

    # 2. Создание конфига спикеров (SpeakerVoiceConfig)
    speaker_configs = []

    # Сортируем спикеров, чтобы "1" всегда получал первый голос, "2" второй и т.д.
    found_speakers.sort()

    for i, speaker_name in enumerate(found_speakers):
        # Логика выбора голоса:
        # 1. Если задан вручную в команде -> берем его
        # 2. Иначе берем из списка по кругу

        if custom_cast and speaker_name in custom_cast:
            voice_name = custom_cast[speaker_name]
        else:
            voice_name = VOICE_NAMES_LIST[i % len(VOICE_NAMES_LIST)]

        speaker_configs.append(
            types.SpeakerVoiceConfig(
                speaker=speaker_name,
                voice_config=types.VoiceConfig(
                    prebuilt_voice_config=types.PrebuiltVoiceConfig(
                        voice_name=voice_name
                    )
                )
            )
        )

    # 3. Запрос к API
    # Используем модель, которая точно поддерживает мультиспикер
    model_id = "gemini-2.5-flash-preview-tts"

    config = types.GenerateContentConfig(
        response_modalities=["AUDIO"],
        speech_config=types.SpeechConfig(
            multi_speaker_voice_config=types.MultiSpeakerVoiceConfig(
                speaker_voice_configs=speaker_configs
            )
        )
    )

    accumulated_data = bytearray()
    mime_type = "audio/wav"

    try:
        # Добавляем инструкцию для модели, чтобы она понимала контекст
        full_prompt = f"TTS the following conversation:\n{script_text}"

        async for chunk in await ai_client.aio.models.generate_content_stream(
                model=model_id,
                contents=full_prompt,
                config=config
        ):
            if chunk.candidates and chunk.candidates[0].content.parts:
                part = chunk.candidates[0].content.parts[0]
                if part.inline_data:
                    accumulated_data.extend(part.inline_data.data)
                    # Обычно приходит audio/wav или audio/x-wav
                    if part.inline_data.mime_type:
                        mime_type = part.inline_data.mime_type

        if not accumulated_data:
            print("DEBUG: No audio data received")
            return None

        # API возвращает PCM WAV без заголовка или с ним, но лучше перестраховаться
        # Если mime_type уже wav, часто заголовок есть, но convert_to_wav добавит RIFF если его нет
        # В твоем примере использовался wave модуль, но struct работает быстрее в async

        # Просто используем нашу функцию, она корректно соберет WAV
        final_wav = convert_to_wav(bytes(accumulated_data), mime_type)

        filename = f"dialog_{int(time.time())}.wav"
        with open(filename, "wb") as f:
            f.write(final_wav)

        return filename

    except Exception as e:
        print(f"MultiSpeaker Error: {e}")
        return None


def convert_to_wav(audio_data: bytes, mime_type: str) -> bytes:
    """Добавляет WAV заголовок к сырым PCM данным."""
    params = parse_audio_mime_type(mime_type)
    channels = 1
    data_size = len(audio_data)
    byte_rate = params["rate"] * channels * (params["bits_per_sample"] // 8)
    block_align = channels * (params["bits_per_sample"] // 8)

    # WAV Header (44 bytes)
    header = struct.pack(
        "<4sI4s4sIHHIIHH4sI",
        b"RIFF",
        36 + data_size,
        b"WAVE",
        b"fmt ",
        16,
        1,
        channels,
        params["rate"],
        byte_rate,
        block_align,
        params["bits_per_sample"],
        b"data",
        data_size
    )
    return header + audio_data

async def create_telegraph_page(title, markdown_text):
    """
    Конвертирует Markdown в HTML и загружает статью на Telegra.ph.
    С механизмом повторных попыток (Retries).
    """

    def _sync_upload():
        # 1. Конвертация (делаем один раз)
        try:
            html_content = markdown.markdown(markdown_text, extensions=['fenced_code', 'tables'])
            html_content = html_content.replace("\n", "<br>")
        except Exception as e:
            return f"Error Markdown Conversion: {e}"

        # 2. Попытки отправки (Retry Loop)
        max_retries = 3
        last_error = None

        for attempt in range(1, max_retries + 1):
            try:
                # Пытаемся создать страницу
                response = telegraph_client.create_page(
                    title=title,
                    html_content=html_content,
                    author_name="Gemini Userbot"
                )
                return response['url']  # Успех!

            except Exception as e:
                print(f"⚠️ Telegraph attempt {attempt} failed: {e}")
                last_error = e
                # Если это не последняя попытка, ждем 2 секунды
                if attempt < max_retries:
                    time.sleep(2)

        # Если все попытки провалились
        return f"Error Telegraph (gave up after {max_retries} tries): {last_error}"

    return await asyncio.to_thread(_sync_upload)


async def transcribe_via_gemini(file_path):
    """
    Загружает файл в Gemini, делает транскрибацию с анализом эмоций и удаляет файл.
    """
    if not ai_client: return None

    try:
        # 1. Загружаем файл в Gemini
        # В новой версии SDK upload делается так:
        file_ref = await ai_client.aio.files.upload(file=file_path)

        # Если это видео, нужно подождать процессинга
        while file_ref.state.name == "PROCESSING":
            await asyncio.sleep(2)
            file_ref = await ai_client.aio.files.get(name=file_ref.name)

        if file_ref.state.name == "FAILED":
            return {"error": "Google File Processing Failed"}

        # 2. Формируем промпт и схему
        prompt = """
        Process the audio/video and generate a detailed transcription.
        Output MUST be in Russian (translate if necessary).

        Requirements:
        1. Identify speakers (Speaker 1, 2 etc).
        2. Timestamps (MM:SS).
        3. Detect primary emotion (Happy, Sad, Angry, Neutral, Excited, Serious).
        4. Provide a summary at the start.
        """

        # Схема ответа (JSON)
        schema = {
            "type": "OBJECT",
            "properties": {
                "summary": {"type": "STRING", "description": "Краткое содержание (summary) на русском."},
                "segments": {
                    "type": "ARRAY",
                    "items": {
                        "type": "OBJECT",
                        "properties": {
                            "time": {"type": "STRING", "description": "MM:SS"},
                            "speaker": {"type": "STRING"},
                            "text": {"type": "STRING", "description": "Текст на русском"},
                            "emotion": {"type": "STRING",
                                        "enum": ["Happy", "Sad", "Angry", "Neutral", "Excited", "Serious"]}
                        },
                        "required": ["time", "speaker", "text", "emotion"]
                    }
                }
            },
            "required": ["summary", "segments"]
        }

        # 3. Запрос к модели
        # Используем Flash для скорости
        response = await ai_client.aio.models.generate_content(
            model="gemini-2.5-flash",
            contents=[
                types.Content(
                    parts=[
                        types.Part.from_uri(file_uri=file_ref.uri, mime_type=file_ref.mime_type),
                        types.Part.from_text(text=prompt)
                    ]
                )
            ],
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                response_schema=schema
            )
        )

        # 4. Удаляем файл из облака (Cleanup)
        await ai_client.aio.files.delete(name=file_ref.name)

        # 5. Возвращаем распаршенный JSON (SDK сам парсит в dict/object если mime_type json)
        # Но для надежности берем text и json.loads, если SDK вернет строку
        try:
            return json.loads(response.text)
        except:
            return response.parsed  # Если SDK уже распарсил

    except Exception as e:
        print(f"Transcribe Error: {e}")
        return {"error": str(e)}

async def convert_wav_to_ogg(wav_path):
    """
    Конвертирует WAV в OGG Opus (формат голосовых Telegram) используя ffmpeg.
    """
    ogg_path = wav_path.replace(".wav", ".ogg")

    # Команда для ffmpeg:
    # -c:a libopus : кодек Opus
    # -b:a 32k     : битрейт (стандарт для голосовых)
    # -vn          : убрать видео (на всякий случай)
    # -y           : перезаписать, если есть
    cmd = [
        "ffmpeg", "-i", wav_path,
        "-c:a", "libopus", "-b:a", "32k", "-vn", "-y",
        ogg_path
    ]

    try:
        # Запускаем процесс асинхронно, чтобы не блокировать бота
        process = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.DEVNULL,  # Скрываем лишний вывод
            stderr=asyncio.subprocess.DEVNULL
        )
        await process.communicate()  # Ждем завершения

        if os.path.exists(ogg_path):
            return ogg_path
        return None
    except Exception as e:
        print(f"FFmpeg Error: {e}")
        return None


def smart_split(text, limit=4000):
    """
    Разбивает длинный текст на куски, стараясь не резать слова.
    Лимит 4000 (с запасом до 4096).
    """
    if len(text) <= limit:
        return [text]

    parts = []
    while text:
        if len(text) <= limit:
            parts.append(text)
            break

        # Ищем перенос строки ближе к концу лимита
        cut = text[:limit].rfind('\n')
        if cut == -1:
            # Если нет переноса, ищем пробел
            cut = text[:limit].rfind(' ')

        if cut == -1:
            # Если вообще нет разделителей, режем жестко
            cut = limit

        parts.append(text[:cut])
        text = text[cut:].lstrip()  # Убираем пробелы в начале следующего куска
    return parts


async def get_message_context(client, message):
    """
    Извлекает текст и медиа из сообщения, на которое ответили (Reply).
    Возвращает кортеж: (текст_контекста, картинка_PIL_или_None)
    """
    reply = message.reply_to_message
    if not reply:
        return "", None

    # 1. Извлекаем текст
    text_context = reply.text or reply.caption or ""

    # Форматируем, чтобы нейросеть понимала, где чьи слова
    if text_context:
        text_context = f"--- Начало пересылаемого сообщения ---\n{text_context}\n--- Конец пересылаемого сообщения ---\n\n"

    # 2. Извлекаем фото (если есть)
    image = None
    if reply.photo:
        try:
            # Скачиваем фото в память (BytesIO)
            photo_io = await client.download_media(reply, in_memory=True)
            if photo_io:
                image = Image.open(photo_io)
        except Exception as e:
            print(f"Error downloading reply photo: {e}")

    return text_context, image

def format_grounding(text, candidates):
    """Добавляет источники (ссылки) к ответу, если они есть"""
    try:
        # Проверяем, есть ли метаданные поиска
        if not candidates or not candidates[0].grounding_metadata:
            return text

        metadata = candidates[0].grounding_metadata
        if not metadata.grounding_chunks:
            return text

        sources_text = "\n\n🌐 **Источники:**"
        unique_links = set()

        for chunk in metadata.grounding_chunks:
            if chunk.web and chunk.web.uri:
                title = chunk.web.title or "Link"
                if chunk.web.uri not in unique_links:
                    sources_text += f"\n🔹 [{title}]({chunk.web.uri})"
                    unique_links.add(chunk.web.uri)

        return text + sources_text
    except Exception:
        return text


def get_ai_config(chat_id=None):
    """Сборка конфигурации: Модель + Поиск + Инструкции"""
    # 1. Получаем модель
    key = SETTINGS.get("model_key", "1")
    model_info = AVAILABLE_MODELS.get(key, AVAILABLE_MODELS["1"])

    # 2. Собираем инструкцию (Глобальная + Локальная)
    sys_instr = SETTINGS.get("sys_global", "")
    if chat_id:
        local_sys = SETTINGS.get("sys_chats", {}).get(str(chat_id), "")
        if local_sys:
            sys_instr = f"{sys_instr}\n\n[Context: {local_sys}]".strip()

    # 3. Настраиваем инструменты (Поиск)
    tools = []
    if model_info["search"]:
        # Включаем Google Search Tool
        tools = [types.Tool(google_search=types.GoogleSearch())]

    config = types.GenerateContentConfig(
        system_instruction=sys_instr if sys_instr else None,
        tools=tools,
        # Для 2.5 Flash можно отключить thinking, если нужно быстрее, но пока оставим дефолт
    )

    return model_info["id"], config


async def ask_gemini_oneshot(contents):
    if not ai_client: return "⚠️ API Key missing."
    model_id, config = get_ai_config()
    try:
        response = await ai_client.aio.models.generate_content(
            model=model_id, contents=contents, config=config
        )
        return format_grounding(response.text, response.candidates)
    except Exception as e:
        return f"Gemini Error ({model_id}): {e}"


async def ask_gemini_chat(chat_id, contents):
    if not ai_client: return "⚠️ API Key missing."
    model_id, config = get_ai_config(chat_id)
    try:
        # ASYNC CREATE (Client.aio.chats.create)
        if chat_id not in ASYNC_CHAT_SESSIONS:
            ASYNC_CHAT_SESSIONS[chat_id] = ai_client.aio.chats.create(
                model=model_id, config=config
            )

        chat = ASYNC_CHAT_SESSIONS[chat_id]
        # ASYNC SEND
        response = await chat.send_message(contents)
        return format_grounding(response.text, response.candidates)
    except Exception as e:
        if chat_id in ASYNC_CHAT_SESSIONS: del ASYNC_CHAT_SESSIONS[chat_id]
        return f"Chat Error: {e}"


async def generate_gemini_tts(text):
    """Генерация голоса через Google Gemini"""
    if not ai_client: return None

    # Получаем настройки
    tts_model_key = SETTINGS.get("tts_model_key", "1")
    model_id = AVAILABLE_TTS_MODELS.get(tts_model_key, "gemini-2.5-pro-preview-tts")

    voice_key = SETTINGS.get("voice_key", "1")
    # Берем ["name"] из словаря
    voice_data = AVAILABLE_VOICES.get(voice_key, AVAILABLE_VOICES["1"])
    voice_name = voice_data["name"]

    # Конфиг для аудио
    config = types.GenerateContentConfig(
        response_modalities=["audio"],
        speech_config=types.SpeechConfig(
            voice_config=types.VoiceConfig(
                prebuilt_voice_config=types.PrebuiltVoiceConfig(
                    voice_name=voice_name
                )
            )
        )
    )

    accumulated_data = bytearray()
    mime_type = "audio/wav"  # Фолбэк

    try:
        # Используем AIO (асинхронный) клиент
        async for chunk in await ai_client.aio.models.generate_content_stream(
                model=model_id,
                contents=text,
                config=config
        ):
            if chunk.candidates and chunk.candidates[0].content.parts:
                part = chunk.candidates[0].content.parts[0]
                if part.inline_data:
                    accumulated_data.extend(part.inline_data.data)
                    mime_type = part.inline_data.mime_type

        if not accumulated_data:
            return None

        # Конвертируем сырой PCM в WAV (добавляем заголовок)
        final_wav = convert_to_wav(bytes(accumulated_data), mime_type)

        filename = f"gemini_voice_{int(time.time())}.wav"
        with open(filename, "wb") as f:
            f.write(final_wav)

        return filename

    except Exception as e:
        print(f"Gemini TTS Error: {e}")
        return None


async def get_sys_info():
    """Получение инфо о системе (для RPi)"""
    try:
        cpu_usage = psutil.cpu_percent(interval=0.1)
        ram = psutil.virtual_memory()
        uptime = datetime.now() - datetime.fromtimestamp(psutil.boot_time())

        # Температура (работает на Linux/RPi)
        temp = "N/A"
        try:
            temps = psutil.sensors_temperatures()
            if 'cpu_thermal' in temps:
                temp = f"{temps['cpu_thermal'][0].current}°C"
        except Exception as e:
            pass

        info = (
            f"🖥 **System Status (RPi):**\n"
            f"🌡 Temp: `{temp}`\n"
            f"🧠 CPU: `{cpu_usage}%`\n"
            f"💾 RAM: `{ram.percent}%` ({ram.used // (1024 * 1024)}MB / {ram.total // (1024 * 1024)}MB)\n"
            f"⏱ Uptime: `{str(uptime).split('.')[0]}`"
        )
        return info
    except Exception as e:
        return f"Sys info error: {e}"


async def generate_freetts(text):
    """
    Генерация голоса через FreeTTS.ru
    Используем ПОЛНУЮ эмуляцию браузера Firefox (Headers + Cookie).
    """
    url_synth = "https://freetts.ru/api/synthesis"
    url_history = "https://freetts.ru/api/history"

    # Твой UID (должен быть свежим из браузера)
    current_uid = "710a7bacbccdad2f8207f2b3a7f921d0"
    voice_id = "NG6FIoMMe4L1"

    # Полная копия твоих хедеров из Firefox
    headers = {
        "Host": "freetts.ru",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:145.0) Gecko/20100101 Firefox/145.0",
        "Accept": "*/*",
        "Accept-Language": "en-US,en;q=0.5",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Referer": "https://freetts.ru/",
        "Origin": "https://freetts.ru",  # Важно для POST запросов
        "DNT": "1",
        "Sec-GPC": "1",
        "Connection": "keep-alive",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "Priority": "u=4",
        # Content-Type aiohttp добавит сам
    }

    # Куки передаем отдельно, aiohttp их правильно отформатирует
    cookies = {"uid": current_uid}

    try:
        async with aiohttp.ClientSession(headers=headers, cookies=cookies) as session:
            # 1. Отправляем запрос на озвучку (POST)
            payload = {
                "text": text,
                "voiceid": voice_id,
                "ext": "mp3"
            }

            async with session.post(url_synth, json=payload) as resp:
                resp_text = await resp.text()

                # Если 666 или другая ошибка
                if resp.status != 200:
                    print(f"DEBUG FreeTTS POST Fail: {resp.status} | {resp_text}")
                    return None, f"HTTP Error: {resp.status}"

                try:
                    data = json.loads(resp_text)
                    # "data": false часто приходит при ошибке 666
                    if data.get("status") == "error":
                        print(f"DEBUG FreeTTS 666/Error: {data}")
                        return None, f"Anti-Bot Error: {data.get('message')}"
                except:
                    pass

            # 2. Ждем и опрашиваем историю (GET)
            for i in range(15):
                await asyncio.sleep(2)

                async with session.get(url_history) as hist_resp:
                    if hist_resp.status != 200:
                        print(f"DEBUG History Fail: {hist_resp.status}")
                        continue

                    hist_data = await hist_resp.json()

                    if hist_data.get("status") == "success" and isinstance(hist_data.get("data"), list):
                        # Ищем задачу (первые 5 записей)
                        for task in hist_data["data"][:5]:
                            # Сравниваем начало текста
                            # text[:15] может не совпасть, если там спецсимволы, пробуем мягкий поиск
                            if text[:10] in task.get("text", ""):

                                if task["status"] == "done":
                                    audio_url = task["url"]
                                    # 3. Скачиваем
                                    async with session.get(audio_url) as audio_resp:
                                        if audio_resp.status == 200:
                                            content = await audio_resp.read()
                                            filename = f"freetts_{int(time.time())}.mp3"
                                            with open(filename, "wb") as f:
                                                f.write(content)
                                            return filename, None

                                elif task["status"] == "error":
                                    return None, "Server Error inside history"

            return None, "Timeout (не найдено в истории)"

    except Exception as e:
        print(f"DEBUG FreeTTS Exception: {e}")
        return None, str(e)

async def download_yandex_track(url: str):
    def _sync_download():
        tracks_paths = []
        try:
            if "track" in url:
                track_id = re.search(r'track/(\d+)', url).group(1)
                tracks = [ym_client.tracks([track_id])[0]]
            elif "album" in url:
                album_id = re.search(r'album/(\d+)', url).group(1)
                album = ym_client.albums_with_tracks(album_id)
                tracks = album.volumes[0]
            else:
                return []

            for track in tracks:
                info = track.get_download_info(get_direct_links=True)
                if not info: continue
                direct_link = info[0].get_direct_link()

                import requests
                track_data = requests.get(direct_link).content

                filename = f"{track.title} - {track.artists[0].name}.mp3"
                # Очистка имени файла от недопустимых символов
                filename = re.sub(r'[\\/*?:"<>|]', "", filename)

                with open(filename, 'wb') as f:
                    f.write(track_data)
                tracks_paths.append(filename)
            return tracks_paths
        except Exception as e:
            logger.error(f"YM Error: {e}")
            return []

    return await asyncio.to_thread(_sync_download)


async def download_video(link: str, quality_mode: int):
    def _sync_dl():
        options = {
            'outtmpl': '%(title)s.%(ext)s',
            'quiet': True,
            'no_warnings': True,
        }
        if quality_mode == 2:
            options.update({
                'format': 'bestaudio/best',
                'postprocessors': [{'key': 'FFmpegExtractAudio', 'preferredcodec': 'mp3'}],
            })
        elif quality_mode == 1:
            options.update({'format': 'bestvideo[height<=480]+bestaudio/best'})
        else:
            options.update({'format': 'bestvideo+bestaudio/best'})

        with yt_dlp.YoutubeDL(options) as ydl:
            info = ydl.extract_info(link, download=True)
            if quality_mode == 2:
                title = info['title']
                # yt-dlp может заменить расширение
                potential_name = f"{title}.mp3"
                # Простая проверка, иногда имя файла сложнее
                return potential_name
            return ydl.prepare_filename(info)

    try:
        return await asyncio.to_thread(_sync_dl)
    except Exception as e:
        logger.error(f"DL Error: {e}")
        return None


async def olx_parser(query: str):
    def _scrape():
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        # Маскируемся, чтобы сервер отдавал контент как человеку
        chrome_options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36")

        driver = webdriver.Chrome(options=chrome_options)

        # Настройка Excel
        wb = Workbook()
        ws = wb.active
        ws.append(['Фото', 'Ссылка', 'Цена', 'Название', 'Дата/Место', 'Состояние'])

        # Красивые ширины колонок
        dims = {'A': 22, 'B': 15, 'C': 20, 'D': 40, 'E': 25, 'F': 15}
        for col, width in dims.items(): ws.column_dimensions[col].width = width

        try:
            url = f"https://www.olx.uz/list/q-{query}/"
            driver.get(url)
            time.sleep(2)  # Ждем инициализации JS

            # Находим карточки через Selenium, чтобы можно было к ним скроллить
            # Используем CSS селектор по data-cy (надежно)
            card_elements = driver.find_elements("css selector", "div[data-cy='l-card']")

            row = 2
            # Берем первые 10
            for card in card_elements[:10]:
                try:
                    # --- SCROLLING (ВАЖНО ДЛЯ КАРТИНОК) ---
                    # Скроллим к элементу, чтобы сработал Lazy Load
                    driver.execute_script("arguments[0].scrollIntoView({behavior: 'instant', block: 'center'});", card)
                    time.sleep(0.5)  # Даем полсекунды на прогрузку картинки

                    # Теперь парсим HTML уже прогруженной карточки
                    card_html = card.get_attribute('outerHTML')
                    soup = BeautifulSoup(card_html, 'lxml')

                    # 1. Текстовые данные
                    title_tag = soup.find("h6") or soup.find("h4")
                    link_tag = soup.find("a")
                    price_tag = soup.find("p", {"data-testid": "ad-price"})

                    if not (title_tag and link_tag): continue

                    title = title_tag.text.strip()
                    price = price_tag.text.strip() if price_tag else "Договорная"

                    href = link_tag.get("href")
                    link = f"https://www.olx.uz{href}" if href.startswith("/") else href

                    # Доп инфо
                    loc_tag = soup.find("p", {"data-testid": "location-date"})
                    loc_date = loc_tag.text.strip() if loc_tag else "-"
                    cond_tag = soup.find("span", title=True)
                    condition = cond_tag['title'] if cond_tag and len(cond_tag['title']) < 20 else "-"

                    # 2. Обработка КАРТИНКИ (HD Quality)
                    img_tag = soup.find("img")
                    if img_tag:
                        # Пытаемся взять src или srcset
                        img_src = img_tag.get("src") or img_tag.get("srcset", "").split()[0]

                        # Бывает, что src всё еще пустой или base64 заглушка
                        if not img_src or "data:image" in img_src:
                            # Пробуем достать из стилей или других атрибутов, но обычно скролл помогает
                            pass

                        if img_src and "http" in img_src:
                            # --- HD FIX ---
                            # Ссылка обычно выглядит как .../image;s=200x200;...
                            # Меняем размер на большой с помощью Regex
                            hd_src = re.sub(r';s=\d+x\d+', ';s=1000x1000', img_src)

                            import requests
                            resp = requests.get(hd_src, timeout=5)
                            if resp.status_code == 200:
                                img = Image.open(BytesIO(resp.content))
                                img.thumbnail((150, 150))  # Для Excel уменьшаем, но источник был качественный

                                path = f"temp_img_{row}.png"
                                img.save(path)
                                excel_img = ExcelImage(path)
                                excel_img.width = 150;
                                excel_img.height = 120
                                ws.add_image(excel_img, f"A{row}")
                                ws.row_dimensions[row].height = 100

                    # Запись в ячейки
                    ws[f"B{row}"] = f'=HYPERLINK("{link}", "Перейти")'
                    ws[f"B{row}"].style = "Hyperlink"
                    ws[f"C{row}"] = price
                    ws[f"D{row}"] = title
                    ws[f"E{row}"] = loc_date
                    ws[f"F{row}"] = condition

                    row += 1
                except Exception as e:
                    print(f"Item parse err: {e}")
                    continue

            fname = f"olx_{query}_{int(datetime.now().timestamp())}.xlsx"
            wb.save(fname)
            return fname
        finally:
            driver.quit()

    return await asyncio.to_thread(_scrape)


async def get_currency(amount, from_cur, to_cur=None):
    url = f"https://v6.exchangerate-api.com/v6/{EXCHANGE_KEY}/latest/{from_cur}"
    async with aiohttp.ClientSession() as session:
        async with session.get(url) as resp:
            data = await resp.json()

    if data.get('result') != 'success':
        return "Ошибка API валют"

    rates = data['conversion_rates']
    flags = {'USD': '🇺🇸', 'EUR': '🇪🇺', 'RUB': '🇷🇺', 'UZS': '🇺🇿'}

    result = f"💰 **{amount} {from_cur}** {flags.get(from_cur, '')}\n\n"
    targets = [to_cur] if to_cur else ['USD', 'UZS', 'RUB']

    for t in targets:
        if t in rates:
            val = round(amount * rates[t], 2)
            result += f"{t} {flags.get(t, '')}: {val:,.2f}\n".replace(",", " ")

    return result


# --- BOT HANDLERS REGISTER ---

def register_handlers(app: Client):
    # 0. HELP COMMAND

    async def edit_or_reply(message, text, **kwargs):
        if message.outgoing:
            await message.edit(text, **kwargs)
            return message  # Возвращаем само сообщение
        else:
            # Если чужое - отвечаем реплаем
            return await message.reply(text, **kwargs)

    @app.on_message(filters.command(["help", "помощь"], prefixes="."))
    async def help_cmd(client, message):
        text = (
            "🤖 **FULL COMMAND LIST**\n\n"

            "🧠 **AI (Нейросети):**\n"
            "• `.ai` [вопрос] — Разовый запрос (без памяти)\n"
            "• `.ait` [тема] — Написать статью в Telegraph\n"
            "• `.chat` [текст] — Диалог с памятью контекста\n"
            "• `.chatt` [текст] — Ответ контекста в Telegraph\n"
            "• `.model` [1-3] — Выбор модели (Google Search)\n"
            "• `.history` — Показать последние 10 сообщений\n"
            "• `.reset` — Сброс памяти чата + Бэкап\n\n"
            "• `.text` / `.stt` — (Reply) Распознать ГС/Видео в текст\n"

            "🎭 **Роли и Инструкции:**\n"
            "• `.sysglobal` [текст] — Глобальная инструкция (для всех чатов)\n"
            "• `.syschat` [текст] — Инструкция для текущего чата\n"
            "• `.syschat -` — Удалить инструкцию чата\n\n"

            "🛠 **Инструменты:**\n"
            "• `.cal` [выражение] — Калькулятор (2+2*2)\n"
            "• `.dl` [1/2] [ссылка] — Скачать (2=mp3, 1=low, 0=best)\n"
            "• `.olx` [запрос] — Парсинг OLX в Excel (с фото)\n"
            "• `.cur` [100] [USD] — Конвертер валют\n"
            "• `.с` [текст] — Удаление пробелов\n"
            "• `.sys` — Статус сервера (RPi)\n\n"
                
            "🔊 **Звук:**\n"
            "• `.say` [текст] — Голосовое сообщение (OGG)\n"
            "• `.saywav` [текст] — Аудиофайл (WAV)\n"
            "• `.voice` [1-6] — Выбрать голос\n"
            "• `.ttsmodel` [1-2] — Выбрать движок\n"
            "• `.sayfree` — FreeTTS (Резерв)\n\n"

            "🤡 **Fun & Spam:**\n"
            "• `.sar` [текст] — СдЕлАтЬ сАрКаЗм\n"
            "• `.spam` [число] [текст] — Спам отдельными сообщениями\n"
            "• `.spam0` [число] [текст] — Спам столбиком (в одном смс)\n"
            "• `.spam1` [число] [текст] — Спам в строчку (слитно)\n"
            "• `.shrek`, `.girl` — ASCII арты\n\n"
            " 🛠 Бот разработан @RevengerNick"
        )
        await edit_or_reply(message, text)

    @app.on_message(filters.command(["s", "c", "с"], prefixes="."))
    async def strip_handler(client, message):
        try:
            # Делим сообщение на ["команда", "остальной_текст"]
            parts = message.text.split(maxsplit=1)

            if len(parts) < 2:
                # Если текста нет, можно ничего не делать или удалить сообщение
                return

            # Берем текст и удаляем ВСЕ пробелы
            clean_text = parts[1].replace(" ", "")

            # Редактируем
            await message.edit(clean_text)
        except Exception as e:
            await message.edit(f"Err: {e}")

    @app.on_message(filters.command(["say", "скажи", "saywav", "sayfile"], prefixes="."))
    async def say_handler(client, message):
        try:
            # Определяем тип отправки по команде
            cmd = message.command[0].lower()
            send_as_file = "wav" in cmd or "file" in cmd

            parts = message.text.split(maxsplit=1)
            user_text = parts[1] if len(parts) > 1 else ""

            # Реплай логика
            reply_text, _ = await get_message_context(client, message)
            if reply_text:
                clean_reply = reply_text.replace("--- Начало пересылаемого сообщения ---\n", "").replace(
                    "\n--- Конец пересылаемого сообщения ---\n\n", "")
                final_text = clean_reply
            else:
                final_text = user_text

            if not final_text:
                return await edit_or_reply(message, "🗣 Текст?")

            if len(final_text) > 4000: final_text = final_text[:4000]

            v_key = SETTINGS.get("voice_key", "1")
            v_name = AVAILABLE_VOICES.get(v_key, AVAILABLE_VOICES["1"])["name"]
            status = await edit_or_reply(message, f"🗣 Gemini ({v_name}) генерирует...")

            # 1. Генерируем WAV (исходник)
            wav_path = await generate_gemini_tts(final_text)

            if wav_path and os.path.exists(wav_path):
                await status.edit("🗣 Отправка...")

                if send_as_file:
                    # --- ВАРИАНТ ФАЙЛ (WAV) ---
                    await client.send_audio(
                        chat_id=message.chat.id,
                        audio=wav_path,
                        performer=f"Gemini {v_name}",
                        title="TTS Audio",
                        caption=f"🗣 **Gemini WAV** ({v_name})"
                    )
                    os.remove(wav_path)
                else:
                    # --- ВАРИАНТ ГОЛОСОВОЕ (OGG) ---
                    # Сначала конвертируем
                    ogg_path = await convert_wav_to_ogg(wav_path)

                    if ogg_path:
                        await client.send_voice(
                            chat_id=message.chat.id,
                            voice=ogg_path,
                            caption=f"🗣 **Gemini Voice** ({v_name})"
                        )
                        os.remove(ogg_path)  # Удаляем OGG
                    else:
                        await status.edit("❌ Ошибка конвертации в OGG.")

                    os.remove(wav_path)  # Удаляем исходный WAV

                if message.outgoing: await message.delete()
                if status != message: await status.delete()
            else:
                await status.edit("❌ Ошибка TTS.")

        except Exception as e:
            await edit_or_reply(message, f"Err: {e}")

    # 2. VOICE SELECTION
    @app.on_message(filters.me & filters.command(["voice", "голос"], prefixes="."))
    async def voice_select_handler(client, message):
        args = message.text.split()
        curr_key = SETTINGS.get("voice_key", "1")

        if len(args) < 2:
            # Формируем списки
            male_list = []
            female_list = []

            for k, v in AVAILABLE_VOICES.items():
                mark = "✅" if k == curr_key else ""
                line = f"`{k}` — **{v['name']}** ({v['desc']}) {mark}"

                if v["gender"] == "M":
                    male_list.append(line)
                else:
                    female_list.append(line)

            text = "🗣 **Голоса (Gemini):**\n\n"
            text += "👨 **МУЖСКИЕ:**\n" + "\n".join(male_list) + "\n\n"
            text += "👩 **ЖЕНСКИЕ:**\n" + "\n".join(female_list)

            text += "\n\nВыбор: `.voice 5`"
            return await message.edit(text)

        choice = args[1]
        if choice in AVAILABLE_VOICES:
            SETTINGS["voice_key"] = choice
            save_settings()
            info = AVAILABLE_VOICES[choice]
            await message.edit(f"✅ Голос установлен: `{info['name']}`\n({info['desc']})")
        else:
            await message.edit("❌ Неверный номер.")

    # 3. TTS MODEL SELECTION (PRO / FLASH)
    @app.on_message(filters.me & filters.command(["ttsmodel", "модельозвучки"], prefixes="."))
    async def tts_model_handler(client, message):
        args = message.text.split()
        curr = SETTINGS.get("tts_model_key", "1")

        if len(args) < 2:
            text = "🎛 **Модель озвучки:**\n\n"
            for k, v in AVAILABLE_TTS_MODELS.items():
                mark = "✅" if k == curr else ""
                text += f"`{k}` — {v} {mark}\n"
            return await message.edit(text)

        c = args[1]
        if c in AVAILABLE_TTS_MODELS:
            SETTINGS["tts_model_key"] = c;
            save_settings()
            await message.edit(f"✅ Модель TTS: `{AVAILABLE_TTS_MODELS[c]}`")
        else:
            await message.edit("❌ Неверно.")

    @app.on_message(filters.command(["ai", "аи"], prefixes="."))
    async def ai_handler(client, message):
        try:
            parts = message.text.split(maxsplit=1)
            prompt = parts[1] if len(parts) > 1 else ""
            reply_txt, reply_img = await get_message_context(client, message)

            if not prompt and not reply_txt and not reply_img:
                return await edit_or_reply(message, "🤖 Введите вопрос.")

            m_name = AVAILABLE_MODELS[SETTINGS.get("model_key", "1")]["name"]
            status = await edit_or_reply(message, f"🤖 Думаю ({m_name})...")

            final = f"{reply_txt}Вопрос: {prompt}" if reply_txt else prompt
            content = [reply_img, final] if reply_img else final

            resp = await ask_gemini_oneshot(content)
            chunks = smart_split(f"**Gemini ({m_name}):**\n\n{resp}")

            await status.edit(chunks[0], disable_web_page_preview=True)
            for c in chunks[1:]: await client.send_message(message.chat.id, c, disable_web_page_preview=True)
        except Exception as e:
            await edit_or_reply(message, f"Err: {e}")

    @app.on_message(filters.command(["dialog", "диалог", "t"], prefixes="."))
    async def dialog_handler(client, message):
        try:
            parts = message.text.split(maxsplit=1)
            raw_input = parts[1] if len(parts) > 1 else ""

            # Поддержка реплая
            reply_text, _ = await get_message_context(client, message)
            if reply_text:
                clean_reply = reply_text.replace("--- Начало пересылаемого сообщения ---\n", "").replace(
                    "\n--- Конец пересылаемого сообщения ---\n\n", "")
                # Если есть реплай, добавляем его к тексту
                raw_input = f"{raw_input}\n{clean_reply}".strip()

            if not raw_input:
                return await edit_or_reply(message,
                                           "🎭 **Диалог**\n"
                                           "Формат:\n"
                                           "`.t`\n"
                                           "`1: Привет!`\n"
                                           "`2: Хай!`\n\n"
                                           "Кастомные голоса:\n"
                                           "`.t 1=Puck 2=Kore`\n"
                                           "`1: ...`"
                                           )

            status = await edit_or_reply(message, "🎭 Распределяю роли...")

            # --- ПАРСИНГ КАСТОМНЫХ ГОЛОСОВ ---
            # Проверяем первую строку на наличие "Имя=Голос"
            lines = raw_input.split("\n")
            first_line = lines[0]
            custom_cast = {}

            # Ищем паттерны вида Name=Voice (например 1=Puck или Batman=Fenrir)
            cast_pairs = re.findall(r"(\w+)=([A-Za-z]+)", first_line)

            if cast_pairs:
                # Если нашли настройки в первой строке
                for name, voice in cast_pairs:
                    # Проверяем, существует ли такой голос в нашем списке (или просто доверяем API)
                    # Лучше проверить, чтобы не получить 400
                    if voice in [v for v in AVAILABLE_VOICES.values()]:  # Проверка по значениям словаря
                        custom_cast[name] = voice
                    # Также проверим по values списка VOICE_NAMES_LIST
                    elif voice in VOICE_NAMES_LIST:
                        custom_cast[name] = voice

                # Удаляем первую строку из скрипта, раз это были настройки
                script = "\n".join(lines[1:])
            else:
                script = raw_input

            # Генерируем
            wav_path = await generate_multispeaker_tts(script, custom_cast)

            if wav_path:
                await status.edit("🎭 Отправка...")
                ogg_path = await convert_wav_to_ogg(wav_path)

                # Формируем описание ролей для caption
                cast_desc = ", ".join([f"{k}={v}" for k, v in custom_cast.items()]) if custom_cast else "Авто-подбор"

                await client.send_voice(
                    chat_id=message.chat.id,
                    voice=ogg_path if ogg_path else wav_path,
                    caption=f"🎭 **Gemini Dialogue**\nroles: {cast_desc}"
                )

                if ogg_path: os.remove(ogg_path)
                os.remove(wav_path)
                if message.outgoing: await message.delete()
                if status != message: await status.delete()
            else:
                await status.edit("❌ Ошибка (проверь имена голосов или формат 'Имя: текст').")

        except Exception as e:
            await edit_or_reply(message, f"Err: {e}")

    # NEW: PODCAST (Auto Script)
    @app.on_message(filters.command(["podcast", "подкаст"], prefixes="."))
    async def podcast_handler(client, message):
        try:
            parts = message.text.split(maxsplit=1)
            topic = parts[1] if len(parts) > 1 else "о погоде"

            status = await edit_or_reply(message, f"🎙 Пишу сценарий: {topic}...")

            # Генерируем сценарий с цифрами 1 и 2 (так проще мапить на М/Ж)
            prompt = (
                f"Напиши короткий диалог (сценарий подкаста) на тему: '{topic}'. "
                "Спикеры: '1' (мужчина, скептик) и '2' (женщина, веселая). "
                "Формат строго:\n1: текст\n2: текст\n"
                "Добавляй эмоции в скобках. Длина: 8-10 реплик. Язык: Русский."
            )

            script_response = await ask_gemini_oneshot(prompt)
            # Чистим Markdown
            script_clean = script_response.replace("**", "").replace("##", "")

            await status.edit(f"🎙 Озвучиваю...\n\n{script_clean[:50]}...")

            # Для подкаста жестко задаем голоса: 1=Puck(М), 2=Aoede(Ж)
            podcast_cast = {"1": "Puck", "2": "Aoede"}

            wav_path = await generate_multispeaker_tts(script_clean, podcast_cast)

            if wav_path:
                ogg_path = await convert_wav_to_ogg(wav_path)
                await client.send_voice(
                    chat_id=message.chat.id,
                    voice=ogg_path if ogg_path else wav_path,
                    caption=f"🎙 **AI Podcast**\nТема: {topic}"
                )
                if ogg_path: os.remove(ogg_path)
                os.remove(wav_path)
                if message.outgoing: await message.delete()
                if status != message: await status.delete()
            else:
                await status.edit("❌ Ошибка озвучки.")

        except Exception as e:
            await edit_or_reply(message, f"Err: {e}")

    # NEW: AI PODCAST GENERATOR (Auto Script + Audio)
    @app.on_message(filters.command(["podcast", "подкаст"], prefixes="."))
    async def podcast_handler(client, message):
        try:
            parts = message.text.split(maxsplit=1)
            topic = parts[1] if len(parts) > 1 else "о технологиях будущего"

            status = await edit_or_reply(message, f"🎙 Пишу сценарий подкаста про: {topic}...")

            # 1. Генерируем сценарий через текстовую модель (2.0 Flash)
            prompt = (
                f"Напиши короткий, живой диалог (сценарий подкаста) на тему: '{topic}'. "
                "Участники: Алекс (скептик) и Ева (оптимист). "
                "Используй формат:\nАлекс: текст\nЕва: текст\n"
                "Добавляй эмоции в скобках, например: (смеясь) Ева: ...\n"
                "Длина: около 10-12 реплик. Язык: Русский."
            )

            script_response = await ask_gemini_oneshot(prompt)
            # Убираем лишнее, оставляем только текст сценария (если там есть вступления)
            # Обычно Gemini Flash слушается хорошо, но можно почистить:
            script_clean = script_response.replace("**", "").replace("##", "")  # Markdown cleaning

            await status.edit(f"🎙 Сценарий готов:\n\n{script_clean[:100]}...\n\nОзвучиваю...")

            # 2. Озвучиваем сгенерированный сценарий
            wav_path = await generate_multispeaker_tts(script_clean)

            if wav_path:
                ogg_path = await convert_wav_to_ogg(wav_path)

                await client.send_voice(
                    chat_id=message.chat.id,
                    voice=ogg_path if ogg_path else wav_path,
                    caption=f"🎙 **AI Podcast**\nТема: {topic}"
                )

                if ogg_path: os.remove(ogg_path)
                os.remove(wav_path)
                if message.outgoing: await message.delete()
                if status != message: await status.delete()
            else:
                await status.edit("❌ Ошибка озвучки.")

        except Exception as e:
            await edit_or_reply(message, f"Err: {e}")

    @app.on_message(filters.command(["text", "текст", "stt"], prefixes="."))
    async def stt_handler(client, message):
        try:
            # Работаем только с реплаем
            reply = message.reply_to_message
            if not reply or not (reply.voice or reply.audio or reply.video or reply.video_note):
                return await edit_or_reply(message, "⚠️ Ответьте на голосовое, аудио или видео сообщение.")

            m_name = AVAILABLE_MODELS[SETTINGS.get("model_key", "1")]["name"]
            status = await edit_or_reply(message, f"👂 {m_name} слушает и скачивает...")

            # 1. Скачиваем файл
            # limit=50*1024*1024 (50MB) чтобы не ждать вечность на RPi, хотя Gemini жует до 2ГБ
            file_path = await client.download_media(reply)

            if not file_path:
                return await status.edit("❌ Ошибка скачивания.")

            # 2. Отправляем в Gemini
            await status.edit("🧠 Распознаю и анализирую...")
            result = await transcribe_via_gemini(file_path)

            # Удаляем локальный файл сразу
            if os.path.exists(file_path):
                os.remove(file_path)

            if not result or "error" in result:
                return await status.edit(f"❌ Ошибка API: {result.get('error', 'Unknown')}")

            # 3. Формируем отчет
            summary = result.get("summary", "Нет описания")
            segments = result.get("segments", [])

            # Заголовок
            output_text = f"📝 **Транскрипция**\n\n📌 **Суть:** {summary}\n\n"

            # Эмодзи для эмоций
            emojis = {
                "Happy": "😄", "Sad": "😔", "Angry": "😡",
                "Neutral": "😐", "Excited": "🤩", "Serious": "🤔"
            }

            # Собираем диалог
            for seg in segments:
                emo = emojis.get(seg.get('emotion'), "🗣")
                line = f"`{seg['time']}` {emo} **{seg['speaker']}:** {seg['text']}\n"
                output_text += line

            # 4. Отправляем (Чат или Telegraph)
            if len(output_text) > 4000:
                await status.edit("📝 Текст длинный, создаю статью...")
                link = await create_telegraph_page("Audio Transcription", output_text)
                await status.edit(f"📝 **Транскрипция готова:**\n📌 **Суть:** {summary}\n\n👉 {link}")
            else:
                await status.edit(output_text)

        except Exception as e:
            await edit_or_reply(message, f"Err: {e}")

    @app.on_message(filters.command(["ait", "аит"], prefixes="."))
    async def ait_handler(client, message):
        try:
            parts = message.text.split(maxsplit=1)
            prompt = parts[1] if len(parts) > 1 else "Analysis"
            reply_txt, reply_img = await get_message_context(client, message)

            m_name = AVAILABLE_MODELS[SETTINGS.get("model_key", "1")]["name"]
            status = await edit_or_reply(message, f"📝 {m_name} пишет статью...")

            final = f"{reply_txt}\nЗадание: {prompt}"
            content = [reply_img, final] if reply_img else final

            resp = await ask_gemini_oneshot(content)
            title = f"AI: {prompt[:30]}..."
            link = await create_telegraph_page(title, resp)

            await status.edit(f"🧠 **Gemini ({m_name}):**\n📄 **Статья:**\n👉 {link}")
        except Exception as e:
            await edit_or_reply(message, f"Err: {e}")

    @app.on_message(filters.command(["cal", "кал", "calc", "счет"], prefixes="."))
    async def calc_handler(client, message):
        try:
            # Получаем выражение
            args = message.text.split(maxsplit=1)
            if len(args) < 2:
                return await edit_or_reply(message, "🔢 Введите выражение: `.cal 2+2`")

            # Убираем пробелы и заменяем некоторые знаки для удобства
            expr = args[1].lower().replace(" ", "")
            expr = expr.replace("х", "*").replace("x", "*")  # Русская и англ Х на умножение
            expr = expr.replace("^", "**")  # Степень
            expr = expr.replace(":", "/")  # Деление
            expr = expr.replace(",", ".")  # Запятая на точку

            # БЕЗОПАСНОСТЬ: Разрешаем только цифры и мат. знаки
            allowed_chars = set("0123456789.+-*/()%**")
            if not set(expr).issubset(allowed_chars):
                return await edit_or_reply(message, "❌ Ошибка: Недопустимые символы.")

            # Считаем
            # eval безопасен здесь, так как мы проверили символы выше
            result = eval(expr, {"__builtins__": None}, {})

            # Форматируем (убираем .0 если число целое)
            if isinstance(result, (int, float)):
                if int(result) == result:
                    result = int(result)
                # Округляем до 4 знаков, если дробь длинная
                else:
                    result = round(result, 4)

            await edit_or_reply(message, f"🔢 **{args[1]}** = `{result}`")

        except ZeroDivisionError:
            await edit_or_reply(message, "❌ Деление на ноль!")
        except Exception as e:
            await edit_or_reply(message, f"❌ Ошибка: {e}")

    @app.on_message(filters.command(["cur", "кон"], prefixes="."))
    async def cur_handler(client, message):
        try:
            a = message.text.split()
            if len(a) < 3: return await edit_or_reply(message, "⚠️ .cur 100 USD")
            res = await get_currency(float(a[1]), a[2].upper(), a[3].upper() if len(a) > 3 else None)
            await edit_or_reply(message, res)
        except:
            await edit_or_reply(message, "Err")

    # 1. AI CHAT (CONTEXT AWARE)
    @app.on_message(filters.command(["chat", "чат"], prefixes="."))
    async def chat_handler(client, message):
        try:
            parts = message.text.split(maxsplit=1)
            prompt = parts[1] if len(parts) > 1 else ""

            # Получаем контекст (если был реплай)
            reply_txt, reply_img = await get_message_context(client, message)

            if not prompt and not reply_txt and not reply_img:
                return await message.edit("💬 Введите текст или ответьте на сообщение.")

            # Получаем имя модели
            m_name = AVAILABLE_MODELS[SETTINGS.get("model_key", "1")]["name"]

            # Статус "Думаю..."
            await edit_or_reply(message, f"💬 {m_name} думает...")

            # Формируем полный запрос для ИИ
            final_prompt = f"{reply_txt}{prompt}"
            content = [reply_img, final_prompt] if reply_img else final_prompt

            # Делаем запрос
            resp = await ask_gemini_chat(message.chat.id, content)
            chunks = smart_split(resp)

            # Формируем красивый заголовок для первого сообщения
            # Если prompt был пустой (только реплай), пишем "Контекст"
            user_header = f"👤 **Вы:** {prompt}" if prompt else "👤 **Контекст реплая**"

            # Собираем первое сообщение
            first_msg = f"{user_header}\n\n🤖 **{m_name}:**\n{chunks[0]}"

            await edit_or_reply(message, first_msg)

            # Если ответ длинный, отправляем остальные куски следом
            for c in chunks[1:]:
                await client.send_message(message.chat.id, c, disable_web_page_preview=True)

        except Exception as e:
            await edit_or_reply(message, f"Err: {e}")

    @app.on_message(filters.me & filters.command(["chatt", "чатт"], prefixes="."))
    async def chatt_handler(client, message):
        try:
            parts = message.text.split(maxsplit=1)
            prompt = parts[1] if len(parts) > 1 else "Продолжай"
            reply_txt, reply_img = await get_message_context(client, message)

            m_name = AVAILABLE_MODELS[SETTINGS.get("model_key", "1")]["name"]
            await message.edit(f"💬📝 {m_name} пишет...")

            final = f"{reply_txt}{prompt}"
            content = [reply_img, final] if reply_img else final

            resp = await ask_gemini_chat(message.chat.id, content)
            link = await create_telegraph_page(f"Context: {prompt[:20]}...", resp)
            await message.edit(f"💬📝 **Ответ (Telegraph):**\n👉 {link}")
        except Exception as e:
            await message.edit(f"Err: {e}")

    @app.on_message(filters.me & filters.command(["history", "история"], prefixes="."))
    async def history_handler(client, message):
        chat_id = message.chat.id
        if chat_id not in ASYNC_CHAT_SESSIONS: return await message.edit("🤷 Нет диалога")
        try:
            chat = ASYNC_CHAT_SESSIONS[chat_id]
            # ASYNC GET HISTORY
            history = chat.get_history()
            text = "📜 **Последние 10 сообщений:**\n\n"
            print (history)
            for i, msg in enumerate(history[-10:], 1):
                role = "👤" if msg.role == "user" else "🤖"
                content = msg.parts[0].text if msg.parts else "[media]"
                text += f"{role}: {content[:60]}...\n"
            await message.edit(text)
        except Exception as e:
            await message.edit(f"Err: {e}")

    @app.on_message(filters.me & filters.command(["reset", "сброс"], prefixes="."))
    async def reset_handler(client, message):
        chat_id = message.chat.id
        if chat_id in ASYNC_CHAT_SESSIONS:
            try:
                chat = ASYNC_CHAT_SESSIONS[chat_id]
                hist = await chat.get_history()

                msgs = []
                for m in hist:
                    msgs.append({'role': m.role, 'content': m.parts[0].text if m.parts else ""})

                fname = f"chat_backup_{chat_id}_{int(time.time())}.json"
                with open(fname, 'w', encoding='utf-8') as f:
                    json.dump(msgs, f, ensure_ascii=False, indent=2)

                del ASYNC_CHAT_SESSIONS[chat_id]
                await message.edit(f"🧹 Очищено.\n💾 Бэкап: `{fname}`")
            except:
                del ASYNC_CHAT_SESSIONS[chat_id]
                await message.edit("🧹 Очищено.")
        else:
            await message.edit("🧹 Пусто.")

    @app.on_message(filters.me & filters.command(["model", "модель"], prefixes="."))
    async def model_handler(client, message):
        args = message.text.split()
        curr = SETTINGS.get("model_key", "1")
        if len(args) < 2:
            t = "🧠 **Models:**\n\n"
            for k, v in AVAILABLE_MODELS.items():
                mark = "✅" if k == curr else ""
                icon = "🔎" if v["search"] else ""
                t += f"`{k}` — {v['name']} {icon} {mark}\n"
            return await message.edit(t + "\nEx: `.model 2`")

        c = args[1]
        if c in AVAILABLE_MODELS:
            SETTINGS["model_key"] = c;
            save_settings();
            ASYNC_CHAT_SESSIONS.clear()
            m = AVAILABLE_MODELS[c]
            await message.edit(f"✅ Set: `{m['name']}`")
        else:
            await message.edit("❌ Invalid.")

    @app.on_message(filters.me & filters.command(["sysglobal"], prefixes="."))
    async def sysg_handler(client, message):
        if len(message.text.split()) == 1:
            return await message.edit(f"🌐 Global:\n`{SETTINGS.get('sys_global', '-')}`")
        SETTINGS["sys_global"] = message.text.split(maxsplit=1)[1];
        save_settings();
        ASYNC_CHAT_SESSIONS.clear()
        await message.edit(f"🌐 Updated:\n`{SETTINGS['sys_global']}`")

    @app.on_message(filters.me & filters.command(["syschat"], prefixes="."))
    async def sysc_handler(client, message):
        cid = str(message.chat.id)
        if len(message.text.split()) == 1:
            return await message.edit(f"💬 Chat:\n`{SETTINGS.get('sys_chats', {}).get(cid, '-')}`")

        instr = message.text.split(maxsplit=1)[1]
        if "sys_chats" not in SETTINGS: SETTINGS["sys_chats"] = {}

        if instr == "-":
            if cid in SETTINGS["sys_chats"]: del SETTINGS["sys_chats"][cid]
            msg = "🗑 Removed."
        else:
            SETTINGS["sys_chats"][cid] = instr
            msg = f"💬 Set:\n`{instr}`"

        save_settings()
        if message.chat.id in ASYNC_CHAT_SESSIONS: del ASYNC_CHAT_SESSIONS[message.chat.id]
        await message.edit(msg)
    # 4. DOWNLOADER
    @app.on_message(filters.me & filters.command(["dl", "скачать", "дл"], prefixes="."))
    async def dl_handler(client, message):
        args = message.text.split()
        if len(args) < 2:
            return await message.edit("❌ Ссылка?")

        url = args[-1]
        mode = 0
        if len(args) > 2 and args[1].isdigit():
            mode = int(args[1])

        await message.edit("📥 Скачиваю на сервер...")
        path = None

        try:
            if "music.yandex" in url:
                paths = await download_yandex_track(url)
                path = paths[0] if paths else None
            else:
                path = await download_video(url, mode)

            if path and os.path.exists(path):
                await message.edit("📤 Загружаю в Telegram...")

                # Функция прогресс бара
                last_update_time = 0

                async def progress(current, total):
                    nonlocal last_update_time
                    # Обновляем не чаще раза в 2 секунды
                    if time.time() - last_update_time > 2:
                        percent = current * 100 / total
                        try:
                            await message.edit(f"📤 Загрузка: {percent:.1f}%")
                            last_update_time = time.time()
                        except:
                            pass

                await client.send_document(
                    message.chat.id,
                    document=path,
                    caption="✅ Готово",
                    progress=progress
                )
                os.remove(path)
                await message.delete()
            else:
                await message.edit("❌ Ошибка скачивания или файл не найден.")
        except Exception as e:
            await message.edit(f"DL Fatal Error: {e}")

    # 5. OLX PARSER
    @app.on_message(filters.me & filters.command(["olx", "олх"], prefixes="."))
    async def olx_handler_func(client, message):
        try:
            parts = message.text.split(maxsplit=1)
            if len(parts) < 2: return await message.edit("Введите запрос.")

            query = parts[1]
            await message.edit(f"🔍 Паршу OLX: {query}...")
            file_path = await olx_parser(query)

            if file_path:
                await client.send_document(message.chat.id, file_path)
                os.remove(file_path)
                # Чистим картинки
                for f in os.listdir():
                    if f.startswith("temp_img_"): os.remove(f)
                await message.delete()
            else:
                await message.edit("Ничего не найдено.")
        except Exception as e:
            await message.edit(f"OLX Err: {e}")

    # 7. SYSTEM INFO (Raspberry Pi)
    @app.on_message(filters.me & filters.command(["sys", "сис"], prefixes="."))
    async def sys_handler(client, message):
        info = await get_sys_info()
        await message.edit(info)



    # 8. SPAM
    @app.on_message(filters.me & filters.command(["spam", "спам"], prefixes="."))
    async def spam_handler(client, message):
        try:
            _, count, text = message.text.split(maxsplit=2)
            count = int(count)
            await message.delete()
            for _ in range(count):
                await client.send_message(message.chat.id, text)
                await asyncio.sleep(0.3)
        except:
            pass

    @app.on_message(filters.me & filters.command(["spam0", "спам0"], prefixes="."))
    async def spam_handler(client, message):
        try:
            _, count, text = message.text.split(maxsplit=2)
            count = int(count)
            await message.delete()
            text_message = ''
            for _ in range(count):
                text_message += f"{text}\n"

            await client.send_message(message.chat.id, text_message)
        except:
            pass

    @app.on_message(filters.me & filters.command(["spam1", "спам1"], prefixes="."))
    async def spam_handler(client, message):
        try:
            _, count, text = message.text.split(maxsplit=2)
            count = int(count)
            await message.delete()
            text_message = ''
            for _ in range(count):
                text_message += f"{text}"

            await client.send_message(message.chat.id, text_message)
        except:
            pass

    # 9. SARCASM
    @app.on_message(filters.me & filters.command(["sar", "сар"], prefixes="."))
    async def sar_handler(client, message):
        try:
            text = message.text.split(maxsplit=1)[1]
            res = "".join([c.upper() if i % 2 == 0 else c.lower() for i, c in enumerate(text)])
            await message.edit(res)
        except:
            pass

    @app.on_message(filters.command(["шрек", "shrek"], prefixes="."))
    async def sar_handler(client, message):
        try:
            mess = """
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⢟⣩⡍⣙⠛⢛⣿⣿⣿⠛⠛⠛⠛⠻⣿⣿⣿⣿
⠙⢿⣿⣿⣿⡿⠿⠛⠛⢛⣧⣿⠇⠄⠂⠄⠄⠄⠘⣿⣿⣿
⣶⣄⣾⣿⢟⣼⠒⢲⡔⣺⣿⣧⠄⠄⣠⠤⢤⡀⠄⠟⠉⣠
⣿⣿⣿⣿⣿⣟⣀⣬⣵⣿⣿⣿⣶⡤⠙⠄⠘⠃⠄⣴⣾⣿
⣿⣿⣿⣿⣿⡿⢻⠿⢿⣿⣿⠿⠋⠁⠄⠂⠉⠒⢘⣿⣿⣿
⣿⣿⣿⣿⡿⣡⣷⣶⣤⣤⣀⡀⠄⠄⠄⠄⠄⠄⠄⣾⣿⣿
⣿⣿⣿⡿⣸⣿⣿⣿⣿⣿⣿⣿⣷⣦⣰⠄⠄⠄⠄⢾⠿⢿
⣾⣿⣿⣿⡟⠉⠉⠈⠉⠉⠉⠉⠉⠄⠄⠄⠑⠄⠄⠐⡇⠄
⣿⣿⣿⡿⠄⠄⠄⠄⠄⠄⠄⠄⠄⠄⠄⠄⠄⠄⠄⢠⡇⠄
⣿⣿⣿⣯⠄⢠⡀⠄⠄⠄⠄⠄⠄⠄⠄⣀⠄⠄⠄⠄⠁⠄
⣿⣿⣿⣯⣧⣬⣿⣤⣐⣂⣄⣀⣠⡴⠖⠈⠄⠄⠄⠄⠄⠄
⣿⣿⣿⣿⣿⣿⣿⣿⣽⣉⡉⠉⠈⠁⠄⠁⠄⠄⠄⠄⡂⠄
⣿⠿⣿⣿⣿⣿⣷⡤⠈⠉⠉⠁⠄⠄⠄⠄⠄⠄⠄⠠⠔⠄
⢿⣷⣿⣿⢿⣿⣿⣷⡦⢤⡀⠄⠄⠄⠄⠄⠄⢐⣠⡿⠁⠄
    """
            await edit_or_reply(message, mess)
        except:
            pass

    @app.on_message(filters.me & filters.command(["девушка", "girl"], prefixes="."))
    async def sar_handler(client, message):
        try:
            mess = """
⠄⠄⣿⣿⣿⣿⠘⡿⢛⣿⣿⣿⣿⣿⣧⢻⣿⣿⠃⠸⣿⣿⣿⠄⠄⠄⠄⠄
⠄⠄⣿⣿⣿⣿⢀⠼⣛⣛⣭⢭⣟⣛⣛⣛⠿⠿⢆⡠⢿⣿⣿⠄⠄⠄⠄⠄
⠄⠄⠸⣿⣿⢣⢶⣟⣿⣖⣿⣷⣻⣮⡿⣽⣿⣻⣖⣶⣤⣭⡉⠄⠄⠄⠄⠄
⠄⠄⠄⢹⠣⣛⣣⣭⣭⣭⣁⡛⠻⢽⣿⣿⣿⣿⢻⣿⣿⣿⣽⡧⡄⠄⠄⠄
⠄⠄⠄⠄⣼⣿⣿⣿⣿⣿⣿⣿⣿⣶⣌⡛⢿⣽⢘⣿⣷⣿⡻⠏⣛⣀⠄⠄
⠄⠄⠄⣼⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣦⠙⡅⣿⠚⣡⣴⣿⣿⣿⡆⠄
⠄⠄⣰⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⠄⣱⣾⣿⣿⣿⣿⣿⣿⠄
⠄⢀⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⢸⣿⣿⣿⣿⣿⣿⣿⣿⠄
⠄⣸⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡿⠣⣿⣿⣿⣿⣿⣿⣿⣿⣿⠄
⠄⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠿⠛⠑⣿⣮⣝⣛⠿⠿⣿⣿⣿⣿⠄
⢠⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣶⠄⠄⠄⠄⣿⣿⣿⣿⣿⣿⣿⣿⣿⡟⠄ 
                """
            await message.edit(mess)
        except:
            pass

    @app.on_message(filters.me & filters.command(["дэвушка", "assgirl"], prefixes="."))
    async def sar_handler(client, message):
        try:
            mess = """
⣿⣿⣿⣿⠛⠛⠉⠄⠁⠄⠄⠉⠛⢿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⡟⠁⠄⠄⠄⠄⠄⠄⠄⠄⠄⠄⣿⣿⣿⣿⣿⣿⣿
⣿⣿⡇⠄⠄⠄⠐⠄⠄⠄⠄⠄⠄⠄⠠⣿⣿⣿⣿⣿⣿
⣿⣿⡇⠄⢀⡀⠠⠃⡐⡀⠠⣶⠄⠄⢀⣿⣿⣿⣿⣿⣿
⣿⣿⣶⠄⠰⣤⣕⣿⣾⡇⠄⢛⠃⠄⢈⣿⣿⣿⣿⣿⣿
⣿⣿⣿⡇⢀⣻⠟⣻⣿⡇⠄⠧⠄⢀⣾⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣟⢸⣻⣭⡙⢄⢀⠄⠄⠄⠈⢹⣯⣿⣿⣿⣿⣿
⣿⣿⣿⣭⣿⣿⣿⣧⢸⠄⠄⠄⠄⠄⠈⢸⣿⣿⣿⣿⣿
⣿⣿⣿⣼⣿⣿⣿⣽⠘⡄⠄⠄⠄⠄⢀⠸⣿⣿⣿⣿⣿
⡿⣿⣳⣿⣿⣿⣿⣿⠄⠓⠦⠤⠤⠤⠼⢸⣿⣿⣿⣿⣿
⡹⣧⣿⣿⣿⠿⣿⣿⣿⣿⣿⣿⣿⢇⣓⣾⣿⣿⣿⣿⣿
⡞⣸⣿⣿⢏⣼⣶⣶⣶⣶⣤⣶⡤⠐⣿⣿⣿⣿⣿⣿⣿
⣯⣽⣛⠅⣾⣿⣿⣿⣿⣿⡽⣿⣧⡸⢿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⡷⠹⠛⠉⠁⠄⠄⠄⠄⠄⠄⠐⠛⠻⣿⣿⣿⣿
⣿⣿⣿⠃⠄⠄⠄⠄⠄⣠⣤⣤⣤⡄⢤⣤⣤⣤⡘⠻⣿
⣿⣿⡟⠄⠄⣀⣤⣶⣿⣿⣿⣿⣿⣿⣆⢻⣿⣿⣿⡎⠝
⣿⡏⠄⢀⣼⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡎⣿⣿⣿⣿⠐
⣿⡏⣲⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⢇⣿⣿⣿⡟⣼
⣿⡠⠜⣿⣿⣿⣿⣟⡛⠿⠿⠿⠿⠟⠃⠾⠿⢟⡋⢶⣿
⣿⣧⣄⠙⢿⣿⣿⣿⣿⣿⣷⣦⡀⢰⣾⣿⣿⡿⢣⣿⣿
⣿⣿⣿⠂⣷⣶⣬⣭⣭⣭⣭⣵⢰⣴⣤⣤⣶⡾⢐⣿⣿
⣿⣿⣿⣷⡘⣿⣿⣿⣿⣿⣿⣿⢸⣿⣿⣿⣿⢃⣼⣿⣿
                    """
            await message.edit(mess)
        except:
            pass

    # DEBUG
    @app.on_message()
    async def debug_monitor(client, message):
        # Проверяем, есть ли текст, если нет - берем подпись, если нет - пишем тип
        text_content = message.text or message.caption or f"[{message.media or 'Service Message'}]"

        # Безопасное получение ID отправителя (в каналах from_user может не быть)
        user_id = message.from_user.id if message.from_user else message.chat.id

        # Заменяем переносы строк на пробелы для лога, чтобы не засорять консоль
        clean_text = text_content.replace("\n", " ")
        print(f"DEBUG: Msg from {user_id}: {clean_text[:50]}...")

        message.continue_propagation()

def create_app(phone: str):
    clean_phone = re.sub(r'\D', '', phone)
    app = Client(
        name=f"sessions/{clean_phone}",
        api_id=API_ID,
        api_hash=API_HASH,
        phone_number=phone
    )
    register_handlers(app)
    return app


# --- AUTHENTICATION ---

async def interactive_auth(app: Client):
    print(f"🔄 Check session: {app.name}")
    try:
        await app.connect()
    except Exception as e:
        print(f"Conn err: {e}")
        return False

    try:
        me = await app.get_me()
        print(f"✅ Active: {me.first_name}")
        await app.disconnect()
        return True
    except Exception:
        print("👤 Login required...")

    try:
        sent_code = await app.send_code(app.phone_number)
    except Exception as e:
        print(f"❌ Send code err: {e}")
        await app.disconnect()
        return False

    while True:
        code = input(f"📩 Code for {app.phone_number}: ").strip()
        try:
            await app.sign_in(app.phone_number, sent_code.phone_code_hash, code)
            break
        except PhoneCodeInvalid:
            print("❌ Invalid code.")
        except PhoneCodeExpired:
            print("❌ Expired.")
            await app.disconnect();
            return False
        except SessionPasswordNeeded as e:
            print("🔐 2FA Enabled.")
            hint = getattr(e, "hint", e.password_hint if hasattr(e, "password_hint") else None)
            if hint: print(f"💡 Hint: {hint}")

            while True:
                password = input("🔑 Password: ").strip()
                try:
                    await app.check_password(password)
                    break
                except PasswordHashInvalid:
                    print("❌ Wrong password.")
            break

    await app.disconnect()
    return True


# --- MAIN ---

async def main():
    if not os.path.exists("sessions"):
        os.makedirs("sessions")

    apps_pool = []
    for phone in PHONES:
        clean = phone.strip()
        if clean: apps_pool.append(create_app(clean))

    if not apps_pool:
        print("❌ PHONES missing.")
        return

    # Auth Loop
    valid_apps = []
    print("\n--- AUTH PHASE ---")
    for app in apps_pool:
        if await interactive_auth(app):
            valid_apps.append(app)

    if not valid_apps:
        print("❌ No valid clients.")
        return

    # Start Loop
    print("\n--- START PHASE ---")
    started_apps = []
    for app in valid_apps:
        try:
            await app.start()
            me = await app.get_me()
            print(f"🟢 Started: {me.first_name}")
            started_apps.append(app)
        except Exception as e:
            print(f"❌ Fail start {app.name}: {e}")

    if not started_apps: return

    print("🤖 Bot Running. Press Ctrl+P -> Ctrl+Q to detach.")
    await idle()

    for app in started_apps:
        if app.is_connected: await app.stop()


if __name__ == "__main__":
    asyncio.run(main())