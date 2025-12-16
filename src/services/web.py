import asyncio
import time
import os
import re
import platform # Для определения ОС
import psutil   # Для системной инфо
import aiohttp
import markdown
import requests
from io import BytesIO
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from datetime import datetime
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from telegraph import Telegraph

from src.config import EXCHANGE_KEY
from src.state import SETTINGS, save_settings

# Инициализация Telegraph
telegraph_client = Telegraph()

try:
    # 1. Пробуем загрузить токен из настроек
    stored_token = SETTINGS.get("telegraph_token")

    if stored_token:
        telegraph_client = Telegraph(access_token=stored_token)
        print("✅ Telegraph: Logged in with saved token.")
    else:
        # 2. Если токена нет, регистрируем новый аккаунт
        print("🆕 Telegraph: Creating new account...")
        telegraph_client.create_account(short_name='GeminiBot')

        # 3. Сохраняем токен
        SETTINGS["telegraph_token"] = telegraph_client.get_access_token()
        save_settings()
        print("✅ Telegraph: Account created and saved.")
except Exception as e:
    print(f"❌ Telegraph Init Error: {e}")

# --- ВАЛЮТНЫЕ НАСТРОЙКИ ---

CURRENCY_ALIASES = {
    'USD': ['usd', 'dollar', 'dollars', 'доллар', 'доллара', 'долларов', 'бакс', 'баксов', '$'],
    'EUR': ['eur', 'euro', 'euros', 'евро', 'еврей', '€'],
    'RUB': ['rub', 'ruble', 'rubles', 'рубль', 'рубля', 'рублей', 'деревянных', '₽'],
    'UZS': ['uzs', 'sum', 'sums', 'som', 'soms', 'сум', 'сума', 'сумов', 'сомов'],
    'KZT': ['kzt', 'tenge', 'тенге', 'тг'],
    'CNY': ['cny', 'yuan', 'юань', 'юаня', 'юаней', '¥'],
    'GBP': ['gbp', 'pound', 'pounds', 'фунт', 'фунтов', 'стерлингов', '£'],
    'JPY': ['jpy', 'yen', 'yens', 'йена', 'йены', 'иена'],
    'BTC': ['btc', 'bitcoin', 'биток', 'биткоин'],
    'ETH': ['eth', 'ethereum', 'эфир'],
    'UAH': ['uah', 'hryvnia', 'гривна', 'гривны', 'гривен'],
    'BYN': ['byn', 'ruble', 'белруб', 'зайчиков'],
    'KRW': ['krw', 'won', 'вон'],
    'TRY': ['try', 'lira', 'лир', 'лира']
}

CURRENCY_FLAGS = {
    'USD': '🇺🇸', 'EUR': '🇪🇺', 'RUB': '🇷🇺', 'UZS': '🇺🇿',
    'GBP': '🇬🇧', 'JPY': '🇯🇵', 'KZT': '🇰🇿', 'CNY': '🇨🇳',
    'UAH': '🇺🇦', 'BYN': '🇧🇾', 'BTC': '₿', 'ETH': 'Ξ',
    'TRY': '🇹🇷', 'KRW': '🇰🇷'
}

# Значки валют для красивого вывода
CURRENCY_SYMBOLS = {
    'USD': '$', 'EUR': '€', 'RUB': '₽', 'UZS': 'сум',
    'GBP': '£', 'JPY': '¥', 'KZT': '₸', 'CNY': '¥',
    'BTC': '₿', 'ETH': 'Ξ', 'KRW': '₩', 'TRY': '₺',
    'UAH': '₴', 'BYN': 'Br'
}

def normalize_currency(raw_input: str) -> str:
    clean = raw_input.lower().strip()
    for code, aliases in CURRENCY_ALIASES.items():
        if clean == code.lower() or clean in aliases:
            return code
    return clean.upper()


def sanitize_html_for_telegraph(html_content):
    """
    Telegra.ph не поддерживает H1 и H2 в теле статьи.
    Заменяем их на H3 и H4.
    """
    html_content = html_content.replace("<h1>", "<h3>").replace("</h1>", "</h3>")
    html_content = html_content.replace("<h2>", "<h4>").replace("</h2>", "</h4>")
    return html_content


async def update_help_page(title, markdown_text):
    """
    Создает ИЛИ Редактирует страницу помощи.
    """

    def _sync_action():
        try:
            # Конвертация
            html_content = markdown.markdown(markdown_text, extensions=['fenced_code', 'tables'])
            html_content = html_content.replace("\n", "<br>")
            # ВАЖНО: Убираем запрещенные теги
            html_content = sanitize_html_for_telegraph(html_content)

            # Проверяем сохраненную страницу
            path = SETTINGS.get("help_page_path")

            # --- ПОПЫТКА РЕДАКТИРОВАНИЯ ---
            if path:
                try:
                    telegraph_client.edit_page(
                        path=path,
                        title=title,
                        html_content=html_content,
                        author_name="Gemini Userbot"
                    )
                    return SETTINGS["help_page_url"]
                except Exception as e:
                    print(f"⚠️ Edit failed (creating new): {e}")

            # --- СОЗДАНИЕ НОВОЙ ---
            response = telegraph_client.create_page(
                title=title,
                html_content=html_content,
                author_name="Gemini Userbot"
            )

            SETTINGS["help_page_path"] = response['path']
            SETTINGS["help_page_url"] = response['url']
            save_settings()

            return response['url']

        except Exception as e:
            return f"Error Telegraph: {e}"

    return await asyncio.to_thread(_sync_action)


async def create_telegraph_page(title, markdown_text):
    """
    Создает НОВУЮ статью (для .ait и .chatt).
    """

    def _sync_upload():
        try:
            html_content = markdown.markdown(markdown_text, extensions=['fenced_code', 'tables'])
            html_content = html_content.replace("\n", "<br>")
            # ВАЖНО: Убираем запрещенные теги
            html_content = sanitize_html_for_telegraph(html_content)

            for attempt in range(3):
                try:
                    response = telegraph_client.create_page(
                        title=title,
                        html_content=html_content,
                        author_name="Gemini Bot"
                    )
                    return response['url']
                except Exception as e:
                    print(f"Telegraph attempt {attempt} error: {e}")
                    time.sleep(2)
            return "Error: Timeout"
        except Exception as e:
            return f"Error: {e}"

    return await asyncio.to_thread(_sync_upload)


async def olx_parser(query: str, max_pages: int = 1, with_images: bool = True):
    """
    Парсит OLX.uz (Явное указание путей для RPi).
    """

    def _scrape():
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36")

        # --- ИСПРАВЛЕНИЕ ДЛЯ RASPBERRY PI ---
        service = None

        # 1. Проверяем стандартный путь (apt install chromium-chromedriver)
        if os.path.exists("/usr/bin/chromedriver"):
            service = Service("/usr/bin/chromedriver")
        # 2. Проверяем альтернативный путь (иногда бывает тут)
        elif os.path.exists("/usr/lib/chromium-browser/chromedriver"):
            service = Service("/usr/lib/chromium-browser/chromedriver")

        # Если нашли драйвер, используем его. Если нет - надеемся на удачу (Selenium Manager)
        try:
            if service:
                driver = webdriver.Chrome(service=service, options=chrome_options)
            else:
                print("⚠️ Driver path not found, trying default...")
                driver = webdriver.Chrome(options=chrome_options)
        except Exception as e:
            print(f"Selenium Driver Critical Error: {e}")
            return None

        wb = Workbook()
        ws = wb.active
        ws.append(['Фото', 'Ссылка', 'Цена', 'Название', 'Дата/Место', 'Состояние', 'Страница'])

        dims = {'A': 22, 'B': 15, 'C': 20, 'D': 40, 'E': 25, 'F': 15, 'G': 10}
        for col, w in dims.items(): ws.column_dimensions[col].width = w

        row = 2

        try:
            for page in range(1, max_pages + 1):
                base_url = f"https://www.olx.uz/list/q-{query}/"
                url = base_url if page == 1 else f"{base_url}?page={page}"

                print(f"📄 Scraping Page {page}: {url}")
                driver.get(url)
                time.sleep(2 if page == 1 else 1.5)

                if "Ничего не найдено" in driver.page_source:
                    break

                cards = driver.find_elements("css selector", "div[data-cy='l-card']")
                if not cards: break

                for card in cards:
                    try:
                        driver.execute_script("arguments[0].scrollIntoView({behavior: 'instant', block: 'center'});",
                                              card)
                        time.sleep(0.5 if with_images else 0.1)

                        soup = BeautifulSoup(card.get_attribute('outerHTML'), 'html.parser')

                        title_tag = soup.find("h6") or soup.find("h4")
                        if not title_tag: continue
                        title = title_tag.text.strip()

                        price_tag = soup.find("p", {"data-testid": "ad-price"})
                        price = price_tag.text.strip() if price_tag else "Договорная"

                        link_tag = soup.find("a")
                        href = link_tag.get("href")
                        link = f"https://www.olx.uz{href}" if href.startswith("/") else href

                        loc_tag = soup.find("p", {"data-testid": "location-date"})
                        loc = loc_tag.text.strip() if loc_tag else "-"

                        cond_tag = soup.find("span", title=True)
                        cond = cond_tag['title'] if cond_tag and cond_tag.has_attr('title') and len(
                            cond_tag['title']) < 30 else "-"

                        if with_images:
                            img_tag = soup.find("img")
                            if img_tag:
                                src = img_tag.get("src") or img_tag.get("srcset", "").split()[0]
                                if src and "http" in src:
                                    hd_src = re.sub(r';s=\d+x\d+', ';s=1000x1000', src)
                                    try:
                                        resp = requests.get(hd_src, timeout=3)
                                        if resp.status_code == 200:
                                            img = Image.open(BytesIO(resp.content))
                                            img.thumbnail((150, 150))

                                            path = f"temp_img_{row}.png"
                                            img.save(path)

                                            excel_img = ExcelImage(path)
                                            excel_img.width = 150
                                            excel_img.height = 120
                                            ws.add_image(excel_img, f"A{row}")
                                            ws.row_dimensions[row].height = 100
                                    except:
                                        pass
                        else:
                            ws[f"A{row}"] = "No Image"

                        ws[f"B{row}"] = f'=HYPERLINK("{link}", "Перейти")'
                        ws[f"B{row}"].style = "Hyperlink"
                        ws[f"C{row}"] = price
                        ws[f"D{row}"] = title
                        ws[f"E{row}"] = loc
                        ws[f"F{row}"] = cond
                        ws[f"G{row}"] = page

                        row += 1
                    except Exception as e:
                        print(f"Card Error: {e}")
                        continue

                if len(cards) < 5: break

            fname = f"olx_{query}_{int(time.time())}.xlsx"
            wb.save(fname)
            return fname
        finally:
            driver.quit()

    return await asyncio.to_thread(_scrape)


async def get_currency(amount, raw_from, raw_to=None):
    """
    Конвертация валют: Красивый и понятный вывод.
    """
    from_cur = normalize_currency(raw_from)
    to_cur = normalize_currency(raw_to) if raw_to else None

    url = f"https://v6.exchangerate-api.com/v6/{EXCHANGE_KEY}/latest/{from_cur}"

    try:
        async with aiohttp.ClientSession() as s:
            async with s.get(url) as r:
                data = await r.json()
    except Exception as e:
        return f"❌ Network Error: {e}"

    if data.get('result') != 'success':
        return f"❌ API Error (Invalid currency: {from_cur})"

    rates = data['conversion_rates']
    flag_from = CURRENCY_FLAGS.get(from_cur, '')

    # Красивое число (10 000.50)
    fmt_amount = f"{amount:,.2f}".replace(",", " ").replace(".", ",")

    # Заголовок сообщения
    res = f"💸 **Конвертация:**\n"
    res += f"{flag_from} **{fmt_amount} {from_cur}** равны:\n\n"

    # Если целевая валюта не задана, берем топ популярных
    if not to_cur:
        targets = ['USD', 'EUR', 'RUB', 'UZS', 'CNY', 'KZT']
    else:
        targets = [to_cur]

    for t in targets:
        # Не конвертируем в саму себя
        if t == from_cur: continue

        if t in rates:
            val = amount * rates[t]
            flag_to = CURRENCY_FLAGS.get(t, '')
            symbol = CURRENCY_SYMBOLS.get(t, '')

            # Форматирование: 1 234.56
            val_str = f"{val:,.2f}".replace(",", " ").replace(".", ",")

            # Строка вида: 🇷🇺 RUB: 9 234,43 ₽
            res += f"{flag_to} {t}: **{val_str} {symbol}**\n"

    # Футер с датой
    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    res += f"\n📅 _Курс на {now}_"

    return res


async def get_sys_info():
    """
    Системная информация (Кроссплатформенная, через psutil).
    Работает и на Windows, и на Raspberry Pi.
    """
    try:
        # Определяем ОС
        sys_name = platform.system()

        # 1. CPU & RAM (работает везде)
        cpu_usage = psutil.cpu_percent(interval=0.1)
        ram = psutil.virtual_memory()

        # 2. Uptime
        uptime_seconds = time.time() - psutil.boot_time()
        m, s = divmod(uptime_seconds, 60)
        h, m = divmod(m, 60)
        d, h = divmod(h, 24)
        uptime_str = f"{int(h)}h {int(m)}m"
        if d > 0: uptime_str = f"{int(d)}d {uptime_str}"

        # 3. Температура (Сложно для Windows, легко для Linux)
        temp = "N/A"
        if sys_name == "Linux":
            try:
                # Пробуем через psutil
                temps = psutil.sensors_temperatures()
                if 'cpu_thermal' in temps:
                    temp = f"{temps['cpu_thermal'][0].current}°C"
                # Фолбэк для RPi (файловый)
                elif os.path.exists("/sys/class/thermal/thermal_zone0/temp"):
                    with open("/sys/class/thermal/thermal_zone0/temp", "r") as f:
                        temp = f"{int(f.read()) / 1000:.1f}°C"
            except:
                pass
        else:
            temp = "N/A (Win)"
        model = SETTINGS.get("model_key", "?")

        return (
            f"🖥 **System Info ({sys_name}):**\n"
            f"🌡 Temp: `{temp}`\n"
            f"🧠 CPU: `{cpu_usage}%`\n"
            f"💾 RAM: `{ram.percent}%`\n"
            f"⏱ Uptime: `{uptime_str}`\n"
            f"🤖 AI Model: `{model}`"
        )
    except Exception as e:
        return f"Sys info error: {e}"